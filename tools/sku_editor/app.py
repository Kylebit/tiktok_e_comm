"""临时工具：对照商品名/主图填写 TikTok 批量编辑表中的商家 SKU（seller_sku 列）。"""

from __future__ import annotations

import argparse
import json
import mimetypes
import shutil
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

COL_NAME = 3
COL_CATEGORY = 2
COL_SKU_ID = 6
COL_VARIATION = 7
COL_SELLER_SKU = 12
COL_MAIN_IMAGE = 18

HERE = Path(__file__).resolve().parent
STATIC = HERE / "static"

_wb_lock = threading.Lock()


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_product_row(ws, row: int) -> bool:
    img = ws.cell(row, COL_MAIN_IMAGE).value
    return isinstance(img, str) and img.startswith("http")


def _row_payload(ws, row: int) -> dict:
    return {
        "row": row,
        "product_name": _cell_str(ws.cell(row, COL_NAME).value),
        "category": _cell_str(ws.cell(row, COL_CATEGORY).value),
        "sku_id": _cell_str(ws.cell(row, COL_SKU_ID).value),
        "variation_value": _cell_str(ws.cell(row, COL_VARIATION).value),
        "seller_sku": _cell_str(ws.cell(row, COL_SELLER_SKU).value),
        "main_image": _cell_str(ws.cell(row, COL_MAIN_IMAGE).value),
    }


class SkuEditor:
    def __init__(self, xlsx_path: Path):
        self.path = xlsx_path.resolve()
        self.backup_path = self.path.with_suffix(self.path.suffix + ".bak")
        if not self.backup_path.exists():
            shutil.copy2(self.path, self.backup_path)
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.product_rows = [
            r for r in range(2, self.ws.max_row + 1) if _is_product_row(self.ws, r)
        ]

    def stats(self) -> dict:
        filled = sum(
            1
            for r in self.product_rows
            if _cell_str(self.ws.cell(r, COL_SELLER_SKU).value)
        )
        return {
            "total": len(self.product_rows),
            "filled": filled,
            "pending": len(self.product_rows) - filled,
            "xlsx": str(self.path),
            "backup": str(self.backup_path),
        }

    def list_products(self, pending_only: bool = False) -> list[dict]:
        items = []
        for r in self.product_rows:
            sku = _cell_str(self.ws.cell(r, COL_SELLER_SKU).value)
            if pending_only and sku:
                continue
            items.append(
                {
                    "row": r,
                    "product_name": _cell_str(self.ws.cell(r, COL_NAME).value),
                    "seller_sku": sku,
                    "main_image": _cell_str(self.ws.cell(r, COL_MAIN_IMAGE).value),
                }
            )
        return items

    def get_product(self, row: int) -> dict | None:
        if row not in self.product_rows:
            return None
        return _row_payload(self.ws, row)

    def set_sku(self, row: int, seller_sku: str) -> dict:
        if row not in self.product_rows:
            raise ValueError(f"行 {row} 不是有效商品行")
        seller_sku = seller_sku.strip()
        with _wb_lock:
            self.ws.cell(row, COL_SELLER_SKU).value = seller_sku or None
            self.wb.save(self.path)
        return _row_payload(self.ws, row)

    def next_row(self, row: int, pending_only: bool = False) -> int | None:
        try:
            idx = self.product_rows.index(row)
        except ValueError:
            return self.product_rows[0] if self.product_rows else None
        for r in self.product_rows[idx + 1 :]:
            if not pending_only or not _cell_str(self.ws.cell(r, COL_SELLER_SKU).value):
                return r
        return None

    def prev_row(self, row: int) -> int | None:
        try:
            idx = self.product_rows.index(row)
        except ValueError:
            return None
        if idx <= 0:
            return None
        return self.product_rows[idx - 1]


def _read_json(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", 0))
    if length <= 0:
        return {}
    raw = handler.rfile.read(length)
    return json.loads(raw.decode("utf-8"))


def make_handler(editor: SkuEditor):
    class Handler(BaseHTTPRequestHandler):
        def log_message(self, fmt, *args):
            print(f"[sku_editor] {self.address_string()} - {fmt % args}")

        def _send_json(self, data, status=200):
            body = json.dumps(data, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _send_file(self, path: Path):
            if not path.is_file():
                self.send_error(404)
                return
            data = path.read_bytes()
            ctype = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)

        def do_GET(self):
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            if path in ("/", "/index.html"):
                return self._send_file(STATIC / "index.html")

            if path == "/api/stats":
                return self._send_json(editor.stats())

            if path == "/api/products":
                pending = qs.get("pending", ["0"])[0] in ("1", "true", "yes")
                return self._send_json({"items": editor.list_products(pending_only=pending)})

            if path.startswith("/api/product/"):
                try:
                    row = int(path.rsplit("/", 1)[-1])
                except ValueError:
                    return self._send_json({"error": "无效行号"}, 400)
                item = editor.get_product(row)
                if not item:
                    return self._send_json({"error": "未找到商品"}, 404)
                return self._send_json(item)

            self.send_error(404)

        def do_POST(self):
            parsed = urlparse(self.path)
            if not parsed.path.startswith("/api/product/"):
                return self.send_error(404)

            try:
                row = int(parsed.path.rsplit("/", 1)[-1])
            except ValueError:
                return self._send_json({"error": "无效行号"}, 400)

            try:
                body = _read_json(self)
                sku = str(body.get("seller_sku", ""))
                item = editor.set_sku(row, sku)
                pending = bool(body.get("pending_only"))
                nxt = editor.next_row(row, pending_only=pending)
                return self._send_json(
                    {"ok": True, "item": item, "next_row": nxt, "stats": editor.stats()}
                )
            except ValueError as e:
                return self._send_json({"error": str(e)}, 400)
            except Exception as e:
                return self._send_json({"error": str(e)}, 500)

    return Handler


def main():
    parser = argparse.ArgumentParser(description="TikTok 商家 SKU 对照填写工具")
    parser.add_argument(
        "--xlsx",
        default=r"c:\Users\Windows11\Desktop\Tiktoksellercenter_batchedit_20260624_all_information_template.xlsx",
        help="批量编辑 Excel 路径",
    )
    parser.add_argument("--port", type=int, default=8766)
    parser.add_argument("--host", default="127.0.0.1")
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        raise SystemExit(f"找不到文件: {xlsx}")

    editor = SkuEditor(xlsx)
    stats = editor.stats()
    print(f"已加载 {stats['total']} 个商品，已填 SKU {stats['filled']} 个")
    print(f"原表备份: {stats['backup']}")
    print(f"打开浏览器: http://{args.host}:{args.port}/")

    server = HTTPServer((args.host, args.port), make_handler(editor))
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已退出（表格已实时保存）")


if __name__ == "__main__":
    main()
